"""
설계VE 상세내용 - VE제안 목록 CSV/XLSX 를 한 번에 FAISS 벡터DB로 넣어두기 위한 스크립트.

예상 헤더 (엑셀 상단, CSV 로 저장되든 XLSX 로 저장되든 동일):

기관명,사업명,제안명,제안일자,채택여부,공종분류,키워드,개선전_건설사업비(백만원),개선전_유지관리비(백만원),개선전_계(백만원),개선후_건설사업비(백만원),개선후_유지관리비(백만원),개선후_계(백만원),절감액(백만원),절감율(%),개선전_성능점수(점),개선전_가치점수(점),개선후_성능점수(점),개선후_가치점수(점)

모든 파일의 컬럼 구성이 동일하다고 가정하고,
엑셀 상단에 있는 여러 제목 행들 중에서
"기관명,사업명,제안명,제안일자" 네 개가 모두 포함된 줄을 자동으로 헤더로 인식합니다.

사용 예 (CSV 또는 XLSX 둘 다 가능):

    cd backend
    python -m app.services.ingest_ve_csv data/ve_proposals.xlsx
"""

from __future__ import annotations

import csv
import sys
from datetime import date
from pathlib import Path
from typing import Iterable, Dict, Any

from openpyxl import load_workbook

from ..core.models import DesignChangeInput
from .vectorstore import add_design_change


REQUIRED_COLUMNS = [
    "기관명",
    "사업명",
    "제안명",
    "제안일자",
]


def _parse_date(value: str) -> date:
    """YYYY-MM-DD / YYYY.MM.DD / YYYY/MM/DD 형태를 date 로 변환.

    제안일자가 '--' 처럼 비어 있는 경우가 많으므로,
    파싱이 안 되면 2000-01-01 로 통일해서 저장한다.
    (실제 날짜는 description 텍스트에 그대로 남아 있음)
    """
    v = str(value).strip()
    if not v or v in {"--", "-", "미정"}:
        return date(2000, 1, 1)

    v = v.replace(".", "-").replace("/", "-")
    # 엑셀에서 '2025-12-09 00:00:00' 처럼 나올 수 있으므로 앞 10자리만 사용
    v = v[:10]
    try:
        return date.fromisoformat(v)
    except ValueError:
        # 이상한 형식은 모두 기본값으로 처리
        return date(2000, 1, 1)


def _build_description(row: Dict[str, Any]) -> str:
    """여러 숫자/메타데이터를 한글 설명 텍스트로 합쳐서 description 으로 사용."""
    def g(key: str) -> str:
        return str(row.get(key, "")).strip()

    lines = [
        f"[VE 제안명] {g('제안명')}",
        f"기관명: {g('기관명')}",
        f"사업명: {g('사업명')}",
        f"제안일자: {g('제안일자')}",
        f"채택여부: {g('채택여부')}",
        f"공종분류: {g('공종분류')}",
        f"키워드: {g('키워드')}",
        "",
        "[생애주기비용(LCC) 절감효과 - 개선전]",
        f"- 건설사업 비용(백만원): {g('개선전_건설사업비(백만원)')}",
        f"- 유지관리 비용(백만원): {g('개선전_유지관리비(백만원)')}",
        f"- 계(백만원): {g('개선전_계(백만원)')}",
        "",
        "[생애주기비용(LCC) 절감효과 - 개선후]",
        f"- 건설사업 비용(백만원): {g('개선후_건설사업비(백만원)')}",
        f"- 유지관리 비용(백만원): {g('개선후_유지관리비(백만원)')}",
        f"- 계(백만원): {g('개선후_계(백만원)')}",
        "",
        f"절감액(백만원): {g('절감액(백만원)')}",
        f"절감율(%): {g('절감율(%)')}",
        "",
        "[가치향상효과 - 개선전]",
        f"- 성능점수(점): {g('개선전_성능점수(점)')}",
        f"- 가치점수(점): {g('개선전_가치점수(점)')}",
        "",
        "[가치향상효과 - 개선후]",
        f"- 성능점수(점): {g('개선후_성능점수(점)')}",
        f"- 가치점수(점): {g('개선후_가치점수(점)')}",
    ]
    return "\n".join(lines)


def _load_csv(path: Path) -> Iterable[Dict[str, Any]]:
    """
    CSV 상단에 제목/기준일자 등이 여러 줄 있을 수 있으므로,
    '기관명,사업명,제안명,제안일자,...' 가 포함된 행을 헤더로 자동 인식한다.
    """
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        raw_reader = csv.reader(f)
        header: list[str] | None = None

        for row in raw_reader:
            cells = [c.strip() for c in row]
            if not any(cells):
                continue
            # 기관명/사업명/제안명/제안일자가 모두 들어 있는 행을 헤더로 사용
            if {"기관명", "사업명", "제안명", "제안일자"}.issubset(set(cells)):
                header = cells
                break

        if header is None:
            raise RuntimeError("CSV에서 헤더 행(기관명,사업명,제안명,제안일자 포함)을 찾지 못했습니다.")

        reader = csv.DictReader(f, fieldnames=header)

        missing = [c for c in REQUIRED_COLUMNS if c not in header]
        if missing:
            raise RuntimeError(f"CSV 헤더에 필수 컬럼이 없습니다: {missing}")

        for row in reader:
            # 제안명이 없는 행은 건너뜀
            if not str(row.get("제안명", "")).strip():
                continue
            yield row
def _load_xlsx(path: Path) -> Iterable[Dict[str, Any]]:
    """
    XLSX 파일에서 헤더/데이터 행을 읽어 Dict 로 변환.
    시트 상단에 제목/기준일자 행이 여러 줄 있을 수 있다고 가정한다.
    """
    wb = load_workbook(filename=path, data_only=True)
    ws = wb.active

    header_row_index: int | None = None
    header: list[str] | None = None

    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = [
            (str(c).strip() if c is not None else "") for c in row
        ]
        if not any(cells):
            continue
        if {"기관명", "사업명", "제안명", "제안일자"}.issubset(set(cells)):
            header_row_index = idx
            header = cells
            break

    if header_row_index is None or header is None:
        raise RuntimeError("XLSX에서 헤더 행(기관명,사업명,제안명,제안일자 포함)을 찾지 못했습니다.")

    missing = [c for c in REQUIRED_COLUMNS if c not in header]
    if missing:
        raise RuntimeError(f"XLSX 헤더에 필수 컬럼이 없습니다: {missing}")

    # 헤더 바로 다음 줄부터 데이터 시작
    for row in ws.iter_rows(min_row=header_row_index + 1, values_only=True):
        values = [
            (str(c).strip() if c is not None else "") for c in row
        ]
        if not any(values):
            continue
        row_dict: Dict[str, Any] = {}
        for col_name, value in zip(header, values):
            if col_name:
                row_dict[col_name] = value

        if not str(row_dict.get("제안명", "")).strip():
            continue
        yield row_dict


def _iter_rows_from_file(path: Path) -> Iterable[Dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _load_csv(path)
    if suffix in {".xlsx", ".xlsm"}:
        return _load_xlsx(path)
    raise RuntimeError(f"지원하지 않는 파일 형식입니다: {suffix}")


def ingest_file(path: Path) -> None:
    """단일 CSV/XLSX 파일을 읽어 벡터DB에 적재."""
    if not path.exists():
        print(f"[ERROR] 파일을 찾을 수 없습니다: {path}")
        return

    count_ok = 0
    count_fail = 0

    for row in _iter_rows_from_file(path):
        try:
            change = DesignChangeInput(
                change_date=_parse_date(row["제안일자"]),
                title=str(row["제안명"]).strip(),
                description=_build_description(row),
                author=None,
                organization=str(row.get("기관명", "")).strip() or None,
                project_name=str(row.get("사업명", "")).strip() or None,
                client=str(row.get("요청발주처", "")).strip() or None,
            )
            add_design_change(change)
            count_ok += 1
        except Exception as e:
            print(f"[WARN] 레코드 변환/저장 실패: {e} / 데이터: {row}")
            count_fail += 1

    print(f"[DONE] {path.name} → 총 {count_ok}건 성공, {count_fail}건 실패")


def ingest_path(target: Path) -> None:
    """
    - 파일 경로가 들어오면 그 파일만 처리
    - 디렉터리 경로가 들어오면 내부의 모든 .csv/.xlsx 파일을 한 번에 처리
    """
    if not target.exists():
        print(f"[ERROR] 경로를 찾을 수 없습니다: {target}")
        return

    if target.is_file():
        ingest_file(target)
        return

    files = sorted(
        [
            p
            for p in target.glob("*")
            if p.suffix.lower() in {".csv", ".xlsx", ".xlsm"}
        ]
    )
    if not files:
        print(f"[WARN] 디렉터리 내에 CSV/XLSX 파일이 없습니다: {target}")
        return

    for file_path in files:
        ingest_file(file_path)


def main(argv: list[str] | None = None) -> None:
    argv = argv or sys.argv[1:]
    if not argv:
        print("사용법:")
        print("  단일 파일: python -m app.services.ingest_ve_csv data/ve_proposals.xlsx")
        print("  디렉터리: python -m app.services.ingest_ve_csv data")
        sys.exit(1)

    file_path = Path(argv[0])
    ingest_path(file_path)


if __name__ == "__main__":
    main()


